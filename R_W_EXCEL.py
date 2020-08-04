# -*-coding:utf-8 -*-
# @Time: 2020/7/21 17:02
# @Author:lemon_su
# @Email:854001151@qq.com
# @File: R_W_EXCEL.py
# @Software: PyCharm

# 用来读取测试数据的
from openpyxl import load_workbook

# wb = load_workbook('lemon_69.xlsx')
# sheet = wb['test_case']
# all_case = [] # 存储所有行的测试用例数据
# for i in range(2,sheet.max_row+1):
#     case = [] # 某一行测试用例数据
#     for j in range(1,sheet.max_column-1):
#         case.append(sheet.cell(row=2,column=j).value)
#         print(sheet.cell(row=2,column=j).value)
#     print(case)
#     all_case.append(case)
# print(all_case)

# 代码优化
def read_data(file_name,sheet_name): # 读取数据的函数
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    all_case = [] # 存储所有行的测试用例数据
    for i in range(2,sheet.max_row+1):
        case = [] # 某一行测试用例数据
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row=i,column=j).value)
            # print(sheet.cell(row=2,column=j).value)
        # print(case)
        all_case.append(case)
    return all_case # 返回所有测试用例数据
    # print(all_case)

def write_data(file_name,sheet_name,row,column,value): # 此函数是写入结果到Excel中
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    # 定位单元格存值 行 列 值
    sheet.cell(row=row,column=column).value = value
    # 保存
    wb.save(file_name)


if __name__ == '__main__':
    all_case = read_data('lemon_69.xlsx','test_case')
    print(all_case)